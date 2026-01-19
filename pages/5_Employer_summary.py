"""
Page 3: Employer Summary
Aggregate cost analysis and census demographics for ICHRA contribution evaluation
"""

import streamlit as st
import pandas as pd
import sys
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from utils import DataFormatter, ContributionComparison, render_feedback_sidebar
from database import get_database_connection


st.set_page_config(page_title="Employer Summary", page_icon="📊", layout="wide")

# Sidebar styling and hero section
st.markdown("""
<style>
    [data-testid="stSidebar"] { background-color: #F0F4FA; }
    [data-testid="stSidebarNav"] a { background-color: transparent !important; }
    [data-testid="stSidebarNav"] a[aria-selected="true"] { background-color: #E8F1FD !important; border-left: 3px solid #0047AB !important; }
    [data-testid="stSidebarNav"] a:hover { background-color: #E8F1FD !important; }
    [data-testid="stSidebar"] button { background-color: #E8F1FD !important; border: 1px solid #B3D4FC !important; color: #0047AB !important; }
    [data-testid="stSidebar"] button:hover { background-color: #B3D4FC !important; border-color: #0047AB !important; }
    [data-testid="stSidebar"] [data-testid="stAlert"] { background-color: #E8F1FD !important; border: 1px solid #B3D4FC !important; color: #003d91 !important; }

    .hero-section {
        background: linear-gradient(135deg, #ffffff 0%, #e8f1fd 100%);
        border-radius: 12px;
        padding: 32px;
        margin-bottom: 24px;
        border-left: 4px solid #0047AB;
    }
    .hero-title {
        font-family: 'Poppins', sans-serif;
        font-size: 28px;
        font-weight: 700;
        color: #0a1628;
        margin-bottom: 8px;
    }
    .hero-subtitle {
        font-size: 16px;
        color: #475569;
        margin: 0;
    }
</style>
""", unsafe_allow_html=True)

# Sidebar: Client name for exports
with st.sidebar:
    st.markdown("**📋 Client Name**")
    if 'client_name' not in st.session_state:
        st.session_state.client_name = ''
    st.text_input(
        "Client name",
        placeholder="Enter client name",
        key="client_name",
        help="Used in export filenames",
        label_visibility="collapsed"
    )


# Initialize session state
if 'db' not in st.session_state:
    st.session_state.db = get_database_connection()

if 'census_df' not in st.session_state:
    st.session_state.census_df = None

if 'dependents_df' not in st.session_state:
    st.session_state.dependents_df = None

if 'contribution_analysis' not in st.session_state:
    st.session_state.contribution_analysis = {}

if 'contribution_settings' not in st.session_state:
    st.session_state.contribution_settings = {'default_percentage': 75}


# Page header
st.markdown("""
<div class="hero-section">
    <div class="hero-title">📊 Employer Summary</div>
    <p class="hero-subtitle">Review aggregate costs, contribution analysis, and census demographics</p>
</div>
""", unsafe_allow_html=True)

# Check prerequisites
if st.session_state.census_df is None:
    st.warning("⚠️ No employee census loaded. Please complete **Census input** first.")
    st.info("👉 Go to **1️⃣ Census input** in the sidebar to upload your census")
    st.stop()

st.markdown("---")

# Get census data
census_df = st.session_state.census_df
dependents_df = st.session_state.dependents_df

# Check if dependents are included
has_dependents = (dependents_df is not None and not dependents_df.empty)

# Check if per-employee contribution data exists
has_individual_contribs = ContributionComparison.has_individual_contributions(census_df)

# ============================================================================
# EXECUTIVE SUMMARY
# ============================================================================

st.subheader("📋 Executive summary")

if has_dependents:
    # Four columns when dependents present
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total employees", len(census_df))

    with col2:
        num_dependents = len(dependents_df)
        st.metric("Total dependents", num_dependents)

    with col3:
        total_covered_lives = len(census_df) + num_dependents
        st.metric("Covered lives", total_covered_lives)

    with col4:
        unique_states = census_df['state'].nunique()
        st.metric("States", unique_states)

else:
    # Three columns for employee-only mode
    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Total employees", len(census_df))

    with col2:
        unique_states = census_df['state'].nunique()
        st.metric("States", unique_states)

    with col3:
        unique_rating_areas = census_df['rating_area_id'].nunique()
        st.metric("Rating areas", unique_rating_areas)

st.markdown("---")

# ============================================================================
# CONTRIBUTION SETTINGS SUMMARY
# ============================================================================

st.subheader("⚙️ Contribution settings")

settings = st.session_state.contribution_settings
contribution_type = settings.get('contribution_type', 'percentage')

if contribution_type == 'class_based':
    strategy_name = settings.get('strategy_name', 'Class-Based')
    st.markdown(f"**Strategy:** {strategy_name}")
    st.markdown(f"**Total monthly:** ${settings.get('total_monthly', 0):,.2f}")
    st.markdown(f"**Total annual:** ${settings.get('total_annual', 0):,.2f}")
    st.markdown(f"**Employees assigned:** {settings.get('employees_assigned', 0)}")

    if settings.get('apply_family_multipliers'):
        st.markdown("**Family multipliers:** Enabled (EE=1.0x, ES=1.5x, EC=1.3x, F=1.8x)")

    # Show tier summary if available
    if settings.get('strategy_applied') == 'age_banded' and settings.get('tiers'):
        st.markdown("**Age tiers:**")
        for tier in settings['tiers']:
            st.markdown(f"- {tier['age_range']}: ${tier['contribution']:.0f}/mo")
    elif settings.get('strategy_applied') == 'location_based' and settings.get('tiers'):
        st.markdown("**Location tiers:**")
        for tier in settings['tiers']:
            st.markdown(f"- {tier['location']}: ${tier['contribution']:.0f}/mo")

else:
    contribution_pct = settings.get('default_percentage', 75)
    st.markdown("**Contribution type:** Percentage")
    st.markdown(f"**Employer contribution:** {contribution_pct}%")

    if 'by_class' in settings and settings['by_class']:
        st.markdown("**By employee class:**")
        for class_name, pct in settings['by_class'].items():
            st.markdown(f"- {class_name}: {pct}%")

st.markdown("---")

# ============================================================================
# ICHRA COST IMPACT - Total premium comparison (apples to apples)
# ============================================================================

if has_individual_contribs:
    from financial_calculator import FinancialSummaryCalculator

    # Get contribution totals
    contrib_totals = ContributionComparison.aggregate_contribution_totals(census_df)
    strategy_results = st.session_state.get('strategy_results', {})
    has_strategy = strategy_results.get('calculated', False)

    if has_strategy:
        result = strategy_results.get('result', {})
        proposed_annual = result.get('total_annual', 0)
        proposed_monthly = result.get('total_monthly', 0)
        employees_covered = result.get('employees_covered', 0)
        strategy_name = result.get('strategy_name', 'Applied Strategy')

        # Current costs (2025)
        current_er_annual = contrib_totals['total_current_er_annual']
        current_ee_annual = contrib_totals['total_current_ee_annual']
        current_total_annual = current_er_annual + current_ee_annual
        current_er_monthly = contrib_totals['total_current_er_monthly']
        current_ee_monthly = contrib_totals['total_current_ee_monthly']
        current_total_monthly = current_er_monthly + current_ee_monthly

        # 2026 Renewal TOTAL premium (from census or financial summary)
        renewal_total_annual = 0
        renewal_total_monthly = 0
        has_renewal_data = False
        _financial_summary = st.session_state.get('financial_summary')
        if _financial_summary and _financial_summary.get('renewal_monthly'):
            renewal_total_monthly = _financial_summary['renewal_monthly']
            renewal_total_annual = renewal_total_monthly * 12
            has_renewal_data = True
        else:
            projected_data = FinancialSummaryCalculator.calculate_projected_2026_total(census_df)
            if projected_data['has_data']:
                renewal_total_monthly = projected_data['total_monthly']
                renewal_total_annual = projected_data['total_annual']
                has_renewal_data = True

        # Calculate ER/EE split percentages from current costs
        er_pct = current_er_monthly / current_total_monthly if current_total_monthly > 0 else 0.60
        ee_pct = current_ee_monthly / current_total_monthly if current_total_monthly > 0 else 0.40

        # PROJECT the 2026 ER contribution using same split ratio
        # This is the KEY metric - what would employer pay at renewal
        projected_er_monthly_2026 = renewal_total_monthly * er_pct
        projected_er_annual_2026 = projected_er_monthly_2026 * 12
        projected_ee_monthly_2026 = renewal_total_monthly * ee_pct
        projected_ee_annual_2026 = projected_ee_monthly_2026 * 12

        # Calculate all comparisons
        # 1. vs Current ER (employer-to-employer, same year baseline)
        delta_vs_current_er = proposed_annual - current_er_annual
        delta_vs_current_er_pct = (delta_vs_current_er / current_er_annual * 100) if current_er_annual > 0 else 0

        # 2. vs Projected Renewal ER - THE PRIMARY SALES COMPARISON
        savings_vs_renewal_er = projected_er_annual_2026 - proposed_annual
        savings_vs_renewal_er_pct = (savings_vs_renewal_er / projected_er_annual_2026 * 100) if projected_er_annual_2026 > 0 else 0

        # 3. vs Renewal Total (for context - apples to oranges but big number)
        savings_vs_renewal_total = renewal_total_annual - proposed_annual
        savings_vs_renewal_total_pct = (savings_vs_renewal_total / renewal_total_annual * 100) if renewal_total_annual > 0 else 0

        # === HEADLINE: EMPLOYER COST COMPARISON ===
        st.subheader("💰 Employer cost comparison")

        # === COST BREAKDOWN DETAILS (moved to top) ===
        st.caption(f"Strategy: {strategy_name} · {employees_covered} employees · ER share: {er_pct*100:.1f}%")

        if has_renewal_data:
            detail_col1, detail_col2, detail_col3 = st.columns(3)

            with detail_col1:
                st.markdown("**Current (2025)**")
                st.write(f"- ER: {DataFormatter.format_currency(current_er_annual)}/yr")
                st.write(f"- EE: {DataFormatter.format_currency(current_ee_annual)}/yr")
                st.write(f"- **Total**: {DataFormatter.format_currency(current_total_annual)}/yr")

            with detail_col2:
                st.markdown("**2026 Renewal (Projected)**")
                st.write(f"- ER: {DataFormatter.format_currency(projected_er_annual_2026)}/yr")
                st.write(f"- EE: {DataFormatter.format_currency(projected_ee_annual_2026)}/yr")
                st.write(f"- **Total**: {DataFormatter.format_currency(renewal_total_annual)}/yr")
                st.caption(f"+{((renewal_total_annual/current_total_annual)-1)*100:.1f}% increase" if current_total_annual > 0 else "")

            with detail_col3:
                st.markdown("**Proposed ICHRA**")
                avg_per_emp = proposed_monthly / employees_covered if employees_covered > 0 else 0
                st.write(f"- ER Budget: {DataFormatter.format_currency(proposed_annual)}/yr")
                st.write(f"- Avg/Employee: {DataFormatter.format_currency(avg_per_emp)}/mo")
                st.write(f"- Employees: {employees_covered}")
        else:
            # No renewal data - show only Current vs Proposed
            detail_col1, detail_col2 = st.columns(2)

            with detail_col1:
                st.markdown("**Current (2025)**")
                st.write(f"- ER: {DataFormatter.format_currency(current_er_annual)}/yr")
                st.write(f"- EE: {DataFormatter.format_currency(current_ee_annual)}/yr")
                st.write(f"- **Total**: {DataFormatter.format_currency(current_total_annual)}/yr")

            with detail_col2:
                st.markdown("**Proposed ICHRA**")
                avg_per_emp = proposed_monthly / employees_covered if employees_covered > 0 else 0
                st.write(f"- ER Budget: {DataFormatter.format_currency(proposed_annual)}/yr")
                st.write(f"- Avg/Employee: {DataFormatter.format_currency(avg_per_emp)}/mo")
                st.write(f"- Employees: {employees_covered}")

            st.caption("💡 *No renewal premium data available. Savings calculated vs current employer costs.*")

        # === EXPORT BUTTONS ===
        st.markdown("")
        st.markdown("---")
        st.markdown("**Export Options**")
        export_col1, export_col2, export_col3 = st.columns([2, 1, 1])

        # Build comprehensive Excel export with multiple tabs
        def generate_comprehensive_excel():
            """Generate Excel workbook with tabs: Employee Detail, Summary & Costs, Breakdowns, Glossary."""
            import io
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            output = io.BytesIO()
            client_name = st.session_state.get('client_name', 'Client')
            aff_impact = result.get('affordability_impact', {})
            after = aff_impact.get('after', {}) if aff_impact else {}
            config = result.get('config', {})
            strategy_type = result.get('strategy_type', '')
            emp_contribs = result.get('employee_contributions', {})

            # Define Tahoma 12pt font style
            tahoma_font = Font(name='Tahoma', size=12)
            tahoma_bold = Font(name='Tahoma', size=12, bold=True)
            header_fill = PatternFill(start_color='E8F1FD', end_color='E8F1FD', fill_type='solid')
            section_fill = PatternFill(start_color='F0F4FA', end_color='F0F4FA', fill_type='solid')

            def apply_tahoma_styling(worksheet):
                """Apply Tahoma 12pt font to all cells in worksheet."""
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.font = tahoma_font

            def style_header_row(worksheet, header_row=1):
                """Style the header row with bold font and fill."""
                for cell in worksheet[header_row]:
                    cell.font = tahoma_bold
                    cell.fill = header_fill

            with pd.ExcelWriter(output, engine='openpyxl') as writer:

                # ============================================================
                # TAB 1: EMPLOYEE DETAIL (moved to first position)
                # ============================================================
                if emp_contribs:
                    emp_rows = []
                    for emp_id, emp_data in emp_contribs.items():
                        # Get census data
                        emp_census = census_df[census_df['employee_id'] == emp_id]
                        if not emp_census.empty:
                            emp_row = emp_census.iloc[0]
                            current_er = emp_row.get('current_er_monthly', 0) if pd.notna(emp_row.get('current_er_monthly')) else 0
                            current_ee = emp_row.get('current_ee_monthly', 0) if pd.notna(emp_row.get('current_ee_monthly')) else 0
                            county = emp_row.get('county', '')
                            zip_code = emp_row.get('zip_code', emp_row.get('zip', emp_row.get('home_zip', '')))
                            monthly_income = emp_row.get('monthly_income', 0) if pd.notna(emp_row.get('monthly_income')) else 0
                        else:
                            current_er = current_ee = 0
                            county = zip_code = ''
                            monthly_income = 0

                        emp_aff = aff_impact.get('employee_affordability', {}).get(emp_id, {}) if aff_impact else {}

                        # Determine affordability method and max EE cost
                        is_fpl_strategy = strategy_type == 'fpl_safe_harbor'
                        has_income = monthly_income > 0

                        if is_fpl_strategy:
                            # FPL Safe Harbor strategy - use FPL threshold
                            fpl_threshold = emp_data.get('fpl_threshold') or result.get('config', {}).get('fpl_threshold')
                            max_ee_cost = fpl_threshold
                            affordability_basis = 'FPL'
                            # If income data available, also calculate income-based threshold for reference
                            income_based_threshold = monthly_income * 0.0996 if has_income else None
                        elif has_income:
                            # Income-based strategy with income data
                            max_ee_cost = monthly_income * 0.0996
                            affordability_basis = 'Income'
                            income_based_threshold = max_ee_cost
                        else:
                            # No income data and not FPL strategy
                            max_ee_cost = None
                            affordability_basis = ''
                            income_based_threshold = None

                        lcsp_ee = emp_data.get('lcsp_ee_rate', 0)
                        ichra_monthly = emp_data.get('monthly_contribution', 0)
                        ee_cost_after_ichra = max(0, lcsp_ee - ichra_monthly) if lcsp_ee else None

                        # Determine tier multiplier for LCSP
                        lcsp_tier = emp_data.get('lcsp_tier_premium', 0)
                        tier_mult = round(lcsp_tier / lcsp_ee, 2) if lcsp_ee and lcsp_ee > 0 else 1.0

                        row_data = {
                            'Employee ID': emp_id,
                            'Name': emp_data.get('name', emp_id),
                            'Age': emp_data.get('age', ''),
                            'State': emp_data.get('state', ''),
                            'County': county,
                            'Rating Area': emp_data.get('rating_area', ''),
                            'ZIP': zip_code,
                            'Family Status': emp_data.get('family_status', 'EE'),
                            'Current ER Monthly': current_er,
                            'Current EE Monthly': current_ee,
                            'Current Total Monthly': current_er + current_ee,
                            'LCSP (Self-Only)': lcsp_ee,
                            'Tier Multiplier': tier_mult,
                            'LCSP (Tier/Family)': lcsp_tier,
                            'Age Ratio (ACA)': emp_data.get('age_ratio', 1.0),
                            'ER Minimum Contribution': emp_data.get('base_contribution', 0),
                            'Family Multiplier': emp_data.get('family_multiplier', 1.0),
                            'ER ICHRA Monthly': ichra_monthly,
                            'ER ICHRA Annual': emp_data.get('annual_contribution', 0),
                            'Monthly Income': monthly_income if has_income else None,
                            'Affordability Basis': affordability_basis,
                            'Max EE Cost (9.96%)': max_ee_cost,
                        }

                        # Add income-based threshold column if using FPL but income data exists
                        if is_fpl_strategy:
                            row_data['Income-Based Threshold'] = income_based_threshold

                        row_data.update({
                            'EE Cost After ICHRA': ee_cost_after_ichra,
                            'Affordable': (
                                'Yes' if (
                                    (is_fpl_strategy and emp_data.get('is_fpl_affordable')) or
                                    emp_aff.get('is_affordable')
                                ) else (
                                    'No' if (
                                        (is_fpl_strategy and emp_data.get('is_fpl_affordable') is False) or
                                        emp_aff.get('is_affordable') is False
                                    ) else ''
                                )
                            ),
                            'Min Needed': emp_aff.get('monthly_contribution') if emp_aff.get('monthly_contribution') else None,
                            'Adjusted': 'Yes' if emp_data.get('adjusted_for_affordability') else '',
                        })

                        emp_rows.append(row_data)

                    emp_df = pd.DataFrame(emp_rows)
                    emp_df.to_excel(writer, sheet_name='Employee Detail', index=False)
                    apply_tahoma_styling(writer.sheets['Employee Detail'])
                    style_header_row(writer.sheets['Employee Detail'])

                # ============================================================
                # TAB 2: SUMMARY & COSTS (combined Summary, Cost Comparison, Savings, IRS Compliance)
                # ============================================================
                summary_rows = []

                # --- REPORT INFO ---
                summary_rows.append({'Section': 'REPORT INFO', 'Item': '', 'Value': '', 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'Report Generated', 'Value': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'), 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'Client Name', 'Value': client_name, 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': '', 'Value': '', 'Notes': ''})

                # --- STRATEGY ---
                summary_rows.append({'Section': 'STRATEGY', 'Item': '', 'Value': '', 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'Strategy Name', 'Value': strategy_name, 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'Strategy Type', 'Value': strategy_type, 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'Employees Covered', 'Value': employees_covered, 'Notes': ''})

                # Strategy-specific parameters
                if strategy_type == 'base_age_curve':
                    summary_rows.append({'Section': '', 'Item': 'Base Age', 'Value': config.get('base_age', 21), 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'Base Contribution', 'Value': config.get('base_contribution', 0), 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'Method', 'Value': 'Base contribution × ACA 3:1 age curve ratio', 'Notes': ''})
                elif strategy_type == 'percentage_lcsp':
                    summary_rows.append({'Section': '', 'Item': 'LCSP Percentage', 'Value': f"{config.get('lcsp_percentage', 0):.1f}%", 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'Method', 'Value': 'Percentage of employee LCSP', 'Notes': ''})
                elif strategy_type == 'fpl_safe_harbor':
                    fpl_threshold = config.get('fpl_threshold', 0) or result.get('config', {}).get('fpl_threshold', 0)
                    summary_rows.append({'Section': '', 'Item': 'FPL Threshold', 'Value': f"${fpl_threshold:,.2f}/mo", 'Notes': '9.96% of Federal Poverty Level'})
                    summary_rows.append({'Section': '', 'Item': 'Method', 'Value': 'FPL Safe Harbor (guaranteed affordable)', 'Notes': ''})

                summary_rows.append({'Section': '', 'Item': 'Family Multipliers', 'Value': 'Yes' if config.get('apply_family_multipliers') else 'No', 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'Affordability Adjusted', 'Value': 'Yes' if result.get('affordability_adjusted') else 'No', 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': '', 'Value': '', 'Notes': ''})

                # --- COST COMPARISON ---
                summary_rows.append({'Section': 'COST COMPARISON', 'Item': '', 'Value': '', 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'Current (2025) - ER', 'Value': current_er_annual, 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'Current (2025) - EE', 'Value': current_ee_annual, 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'Current (2025) - Total', 'Value': current_total_annual, 'Notes': ''})

                if has_renewal_data:
                    summary_rows.append({'Section': '', 'Item': '2026 Renewal - ER', 'Value': projected_er_annual_2026, 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': '2026 Renewal - EE', 'Value': projected_ee_annual_2026, 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': '2026 Renewal - Total', 'Value': renewal_total_annual, 'Notes': ''})

                summary_rows.append({'Section': '', 'Item': 'ICHRA Budget (ER)', 'Value': proposed_annual, 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': 'ICHRA Budget (EE)', 'Value': 'Varies by plan selection', 'Notes': ''})
                summary_rows.append({'Section': '', 'Item': '', 'Value': '', 'Notes': ''})

                # --- SAVINGS ---
                summary_rows.append({'Section': 'SAVINGS ANALYSIS', 'Item': '', 'Value': '', 'Notes': ''})
                savings_vs_current = current_total_annual - proposed_annual
                savings_vs_current_pct = (savings_vs_current / current_total_annual * 100) if current_total_annual > 0 else 0

                if has_renewal_data:
                    ichra_70 = proposed_annual * 0.70
                    savings_70 = renewal_total_annual - ichra_70
                    savings_70_pct = (savings_70 / renewal_total_annual * 100) if renewal_total_annual > 0 else 0

                    summary_rows.append({'Section': '', 'Item': 'vs 2026 Renewal', 'Value': savings_vs_renewal_total, 'Notes': f'{savings_vs_renewal_total_pct:.1f}%'})
                    summary_rows.append({'Section': '', 'Item': 'vs 2026 Renewal (70% take rate)', 'Value': savings_70, 'Notes': f'{savings_70_pct:.1f}%'})

                summary_rows.append({'Section': '', 'Item': 'vs Current Plan (2025)', 'Value': savings_vs_current, 'Notes': f'{savings_vs_current_pct:.1f}%'})
                summary_rows.append({'Section': '', 'Item': '', 'Value': '', 'Notes': ''})

                # --- IRS COMPLIANCE ---
                summary_rows.append({'Section': 'IRS COMPLIANCE', 'Item': '', 'Value': '', 'Notes': ''})

                if strategy_type == 'fpl_safe_harbor':
                    summary_rows.append({'Section': '', 'Item': 'Affordability Method', 'Value': 'FPL Safe Harbor', 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'IRS Threshold', 'Value': f"${fpl_threshold:,.2f}/mo", 'Notes': '9.96% of Federal Poverty Level'})
                    summary_rows.append({'Section': '', 'Item': 'Employees Affordable', 'Value': employees_covered, 'Notes': '100% (guaranteed)'})
                    summary_rows.append({'Section': '', 'Item': 'Status', 'Value': 'Compliant', 'Notes': 'FPL Safe Harbor guarantees affordability'})
                elif after:
                    compliance_status = 'Compliant (Adjusted)' if result.get('affordability_adjusted') else (
                        'Action Needed' if after.get('total_gap', 0) > 0 else 'Compliant'
                    )
                    summary_rows.append({'Section': '', 'Item': 'Affordability Method', 'Value': 'Income-Based', 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'IRS Threshold', 'Value': '9.96% of household income', 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'Employees with Income Data', 'Value': after.get('employees_analyzed', 0), 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'Employees Meeting Threshold', 'Value': after.get('affordable_count', 0), 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'Compliance Rate', 'Value': f"{after.get('affordable_pct', 0):.1f}%", 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'Annual Gap to 100%', 'Value': after.get('total_gap', 0), 'Notes': ''})
                    summary_rows.append({'Section': '', 'Item': 'Status', 'Value': compliance_status, 'Notes': ''})

                summary_df = pd.DataFrame(summary_rows)
                summary_df.to_excel(writer, sheet_name='Summary & Costs', index=False)
                apply_tahoma_styling(writer.sheets['Summary & Costs'])
                style_header_row(writer.sheets['Summary & Costs'])

                # Highlight section headers
                ws = writer.sheets['Summary & Costs']
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value and row[0].value.isupper():
                        for cell in row:
                            cell.font = tahoma_bold
                            cell.fill = section_fill

                # ============================================================
                # TAB 3: BREAKDOWNS (combined By Age Tier and By Family Status)
                # ============================================================
                breakdown_rows = []

                # --- BY AGE TIER ---
                by_age = result.get('by_age_tier', {})
                if by_age:
                    breakdown_rows.append({'Category': 'BY AGE TIER', 'Group': '', 'Employee Count': '', 'Total Monthly': '', 'Average Monthly': '', 'Total Annual': ''})
                    for tier, data in sorted(by_age.items()):
                        avg = data['total_monthly'] / data['count'] if data['count'] > 0 else 0
                        breakdown_rows.append({
                            'Category': '',
                            'Group': tier,
                            'Employee Count': data['count'],
                            'Total Monthly': data['total_monthly'],
                            'Average Monthly': avg,
                            'Total Annual': data['total_monthly'] * 12,
                        })
                    breakdown_rows.append({'Category': '', 'Group': '', 'Employee Count': '', 'Total Monthly': '', 'Average Monthly': '', 'Total Annual': ''})

                # --- BY FAMILY STATUS ---
                by_fs = result.get('by_family_status', {})
                if by_fs:
                    breakdown_rows.append({'Category': 'BY FAMILY STATUS', 'Group': '', 'Employee Count': '', 'Total Monthly': '', 'Average Monthly': '', 'Total Annual': ''})
                    for fs, data in sorted(by_fs.items()):
                        avg = data['total_monthly'] / data['count'] if data['count'] > 0 else 0
                        breakdown_rows.append({
                            'Category': '',
                            'Group': fs,
                            'Employee Count': data['count'],
                            'Total Monthly': data['total_monthly'],
                            'Average Monthly': avg,
                            'Total Annual': data['total_monthly'] * 12,
                        })

                if breakdown_rows:
                    breakdown_df = pd.DataFrame(breakdown_rows)
                    breakdown_df.to_excel(writer, sheet_name='Breakdowns', index=False)
                    apply_tahoma_styling(writer.sheets['Breakdowns'])
                    style_header_row(writer.sheets['Breakdowns'])

                    # Highlight section headers
                    ws = writer.sheets['Breakdowns']
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        if row[0].value and row[0].value.startswith('BY '):
                            for cell in row:
                                cell.font = tahoma_bold
                                cell.fill = section_fill

                # ============================================================
                # TAB 4: LCSP PLANS (employee LCSP details with plan info)
                # ============================================================
                if emp_contribs:
                    # Query to get LCSP plan details for each employee
                    lcsp_plan_rows = []

                    # Get unique state/rating_area/age combinations to batch query
                    lcsp_lookups = {}
                    for emp_id, emp_data in emp_contribs.items():
                        state = emp_data.get('state', '')
                        rating_area = emp_data.get('rating_area', '')
                        age = emp_data.get('age', 21)
                        lcsp_ee_rate = emp_data.get('lcsp_ee_rate', 0)

                        if state and rating_area:
                            key = (state, rating_area, age)
                            if key not in lcsp_lookups:
                                lcsp_lookups[key] = {'rate': lcsp_ee_rate, 'employees': []}
                            lcsp_lookups[key]['employees'].append(emp_id)

                    # Query LCSP plan details for each unique combination
                    lcsp_plan_details = {}
                    try:
                        db = st.session_state.get('db')
                        if db:
                            for (state, rating_area, age), lookup_data in lcsp_lookups.items():
                                # Extract numeric rating area
                                if isinstance(rating_area, str):
                                    import re
                                    ra_match = re.search(r'(\d+)', rating_area)
                                    ra_numeric = int(ra_match.group(1)) if ra_match else 1
                                else:
                                    ra_numeric = int(rating_area) if rating_area else 1

                                # Query for LCSP plan details
                                lcsp_query = """
                                SELECT
                                    p.hios_plan_id as plan_id,
                                    p.plan_marketing_name,
                                    i.issr_lgl_name as issuer_name,
                                    v.plan_brochure,
                                    v.url_for_summary_of_benefits_and_coverage as sbc_url,
                                    br.individual_rate::numeric as premium
                                FROM rbis_insurance_plan_base_rates_20251019202724 br
                                JOIN rbis_insurance_plan_20251019202724 p ON br.plan_id = p.hios_plan_id
                                JOIN rbis_insurance_plan_variant_20251019202724 v ON p.hios_plan_id = v.hios_plan_id
                                LEFT JOIN "HIOS_issuers_pivoted" i ON LEFT(p.hios_plan_id, 5) = i.hios_issuer_id
                                WHERE p.level_of_coverage = 'Silver'
                                    AND p.market_coverage = 'Individual'
                                    AND v.csr_variation_type = 'Exchange variant (no CSR)'
                                    AND SUBSTRING(p.hios_plan_id, 6, 2) = %s
                                    AND br.rating_area_id ~ '^Rating Area [0-9]+$'
                                    AND (REGEXP_REPLACE(br.rating_area_id, '[^0-9]', '', 'g'))::integer = %s
                                    AND br.age = %s
                                ORDER BY br.individual_rate::numeric ASC
                                LIMIT 1
                                """
                                try:
                                    result_df = pd.read_sql(lcsp_query, db.engine, params=(state, ra_numeric, str(age)))
                                    if not result_df.empty:
                                        row = result_df.iloc[0]
                                        for emp_id in lookup_data['employees']:
                                            lcsp_plan_details[emp_id] = {
                                                'plan_id': row.get('plan_id', ''),
                                                'plan_marketing_name': row.get('plan_marketing_name', ''),
                                                'issuer_name': row.get('issuer_name', ''),
                                                'plan_brochure': row.get('plan_brochure', ''),
                                                'sbc_url': row.get('sbc_url', ''),
                                            }
                                except Exception as e:
                                    pass  # Skip on error, leave empty
                    except Exception as e:
                        pass  # Skip LCSP plan lookup on error

                    # Build LCSP Plans rows
                    for emp_id, emp_data in emp_contribs.items():
                        # Get census data for county and ZIP
                        emp_census = census_df[census_df['employee_id'] == emp_id]
                        if not emp_census.empty:
                            emp_row = emp_census.iloc[0]
                            county = emp_row.get('county', '')
                            zip_code = emp_row.get('zip_code', emp_row.get('zip', emp_row.get('home_zip', '')))
                        else:
                            county = ''
                            zip_code = ''

                        plan_details = lcsp_plan_details.get(emp_id, {})

                        lcsp_plan_rows.append({
                            'Employee ID': emp_id,
                            'Name': emp_data.get('name', emp_id),
                            'Age': emp_data.get('age', ''),
                            'State': emp_data.get('state', ''),
                            'County': county,
                            'Rating Area': emp_data.get('rating_area', ''),
                            'ZIP': zip_code,
                            'Family Status': emp_data.get('family_status', 'EE'),
                            'LCSP Premium': emp_data.get('lcsp_ee_rate', 0),
                            'Plan Name': plan_details.get('plan_marketing_name', ''),
                            'Plan ID': plan_details.get('plan_id', ''),
                            'Carrier': plan_details.get('issuer_name', ''),
                            'Plan Brochure': plan_details.get('plan_brochure', ''),
                            'SBC Link': plan_details.get('sbc_url', ''),
                        })

                    if lcsp_plan_rows:
                        lcsp_df = pd.DataFrame(lcsp_plan_rows)
                        lcsp_df.to_excel(writer, sheet_name='LCSP Plans', index=False)
                        apply_tahoma_styling(writer.sheets['LCSP Plans'])
                        style_header_row(writer.sheets['LCSP Plans'])

                # ============================================================
                # TAB 5: GLOSSARY
                # ============================================================
                glossary_data = {
                    'Term': [
                        'ICHRA',
                        'LCSP',
                        'LCSP (Self-Only)',
                        'LCSP (Tier/Family)',
                        'LCSP Premium',
                        'Plan ID',
                        'Carrier',
                        'Plan Brochure',
                        'SBC Link',
                        'Tier Multiplier',
                        'Rating Area',
                        'ACA Age Curve',
                        'Age Ratio',
                        'Family Multiplier',
                        'ER Minimum Contribution',
                        'ER ICHRA Monthly',
                        'ER ICHRA Annual',
                        'Affordability Basis',
                        '9.96% Threshold (Income)',
                        'FPL Safe Harbor',
                    ],
                    'Definition': [
                        'Individual Coverage Health Reimbursement Arrangement - employer-funded allowance for employees to buy their own health insurance',
                        'Lowest Cost Silver Plan - cheapest silver-tier marketplace plan in an employees rating area',
                        'LCSP premium for employee-only (self) coverage based on employee age',
                        'Estimated LCSP premium for the employees coverage tier (Self-Only × Tier Multiplier)',
                        'Monthly premium for the Lowest Cost Silver Plan based on employee age and rating area',
                        'HIOS Plan ID - unique identifier for the marketplace plan (format: XXXXX-CC-NNNNNNNN-NN)',
                        'Insurance company (issuer) offering the plan',
                        'URL to the plan brochure/marketing materials from the carrier',
                        'URL to the Summary of Benefits and Coverage (SBC) document - standardized plan comparison document',
                        'Multiplier applied to Self-Only LCSP based on family status: EE=1.0, ES=1.5, EC=1.3, F=1.8',
                        'Geographic zone used by insurers to set premiums (defined by state)',
                        'Federal 3:1 age rating curve - premiums can vary up to 3x based on age (21-64)',
                        'Multiplier from ACA age curve (age 21 = 1.0, age 64 = 3.0)',
                        'Multiplier for employees with dependents applied to ICHRA contribution',
                        'Base employer contribution before family multipliers are applied',
                        'Total monthly employer ICHRA contribution (after all multipliers)',
                        'Total annual employer ICHRA contribution (ER ICHRA Monthly × 12)',
                        'Method used to determine affordability: "FPL" (Federal Poverty Level) or "Income" (household income)',
                        '2026 IRS affordability threshold - employee LCSP cost cannot exceed 9.96% of household income',
                        'IRS safe harbor: if employee LCSP cost ≤ 9.96% of FPL (~$128/mo), ICHRA is deemed affordable for all employees regardless of income',
                    ],
                }
                glossary_df = pd.DataFrame(glossary_data)
                glossary_df.to_excel(writer, sheet_name='Glossary', index=False)
                apply_tahoma_styling(writer.sheets['Glossary'])
                style_header_row(writer.sheets['Glossary'])

            output.seek(0)
            return output.getvalue()

        with export_col2:
            # Excel Export
            excel_data = generate_comprehensive_excel()
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            client_name = st.session_state.get('client_name', '').strip()
            if client_name:
                safe_name = client_name.replace(' ', '_').replace('/', '-')
                excel_filename = f"ichra_analysis_{safe_name}_{timestamp}.xlsx"
            else:
                excel_filename = f"ichra_analysis_{timestamp}.xlsx"
            st.download_button(
                label="📊 Export Excel",
                data=excel_data,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary",
                key="employer_summary_excel_export"
            )

        with export_col3:
            if st.button("📄 Export PDF", type="secondary", key="employer_summary_pdf_export"):
                with st.spinner("Generating PDF..."):
                    try:
                        from pdf_employer_summary_renderer import generate_employer_summary_pdf, build_employer_summary_data

                        # Build renewal data dict
                        renewal_data = {
                            'renewal_total_annual': renewal_total_annual,
                            'projected_er_annual': projected_er_annual_2026,
                            'projected_ee_annual': projected_ee_annual_2026
                        }

                        # Get client name
                        client_name = st.session_state.get('client_name', 'Client')

                        # Get affordability impact if available
                        affordability_impact = strategy_results.get('result', {}).get('affordability_impact')

                        # Generate PDF
                        pdf_buffer = generate_employer_summary_pdf(
                            strategy_results=strategy_results,
                            contrib_totals=contrib_totals,
                            renewal_data=renewal_data,
                            client_name=client_name if client_name else 'Client',
                            affordability_impact=affordability_impact
                        )

                        # Build filename with client name and timestamp
                        pdf_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        if client_name:
                            safe_name = client_name.replace(' ', '_').replace('/', '-')
                            pdf_filename = f"employer_summary_{safe_name}_{pdf_timestamp}.pdf"
                        else:
                            pdf_filename = f"employer_summary_{pdf_timestamp}.pdf"

                        # Offer download
                        st.download_button(
                            label="⬇️ Download Employer Summary PDF",
                            data=pdf_buffer,
                            file_name=pdf_filename,
                            mime="application/pdf",
                            type="primary",
                            key="employer_summary_pdf_download"
                        )
                    except Exception as e:
                        st.error(f"Error generating PDF: {str(e)}")
                        import logging
                        logging.exception("Employer summary PDF generation error")

    else:
        # No strategy applied yet
        st.subheader("💰 Employer cost impact")
        st.info("""
        **Configure a contribution strategy to see cost comparison**

        Go to **Contribution evaluation** → Use the strategy modeler → Click **"Use this strategy"**
        """)

    # Detailed employee breakdown
    if st.session_state.contribution_analysis:
        with st.expander("📋 View employee-level comparison", expanded=False):
            comparison_rows = []

            for emp_id, analysis in st.session_state.contribution_analysis.items():
                emp_data = census_df[census_df['employee_id'] == emp_id]
                if emp_data.empty:
                    continue

                emp = emp_data.iloc[0]
                current_ee = emp.get('current_ee_monthly')
                current_er = emp.get('current_er_monthly')

                ichra_data = analysis.get('ichra_analysis', {})
                proposed_ee = ichra_data.get('employee_cost', 0)
                proposed_er = ichra_data.get('employer_contribution', 0)
                lcsp_premium = ichra_data.get('monthly_premium', 0)

                # Calculate changes and totals
                ee_change = None
                er_change = None
                total_change = None

                if pd.notna(current_ee):
                    ee_change = proposed_ee - current_ee
                if pd.notna(current_er):
                    er_change = proposed_er - current_er

                # Calculate totals
                current_total = None
                if pd.notna(current_ee) and pd.notna(current_er):
                    current_total = current_ee + current_er
                    total_change = lcsp_premium - current_total

                comparison_rows.append({
                    'Employee ID': emp_id,
                    'Age': int(emp.get('age', 0)) if pd.notna(emp.get('age')) else 'N/A',
                    'State': emp.get('state', 'N/A'),
                    'County': emp.get('county', 'N/A'),
                    'Rating Area': emp.get('rating_area_id', 'N/A'),
                    'Family Status': emp.get('family_status', 'EE'),
                    'LCSP Plan ID': ichra_data.get('plan_id', 'N/A'),
                    'LCSP Plan Name': ichra_data.get('plan_name', 'N/A'),
                    'LCSP Premium': DataFormatter.format_currency(lcsp_premium),
                    'Current EE': DataFormatter.format_currency(current_ee) if pd.notna(current_ee) else 'N/A',
                    'Current ER': DataFormatter.format_currency(current_er) if pd.notna(current_er) else 'N/A',
                    'Current Total': DataFormatter.format_currency(current_total) if current_total is not None else 'N/A',
                    'ICHRA ER': DataFormatter.format_currency(proposed_er),
                    'ICHRA EE': DataFormatter.format_currency(proposed_ee),
                    'ER Change': DataFormatter.format_currency(er_change, include_sign=True) if er_change is not None else 'N/A',
                    'EE Change': DataFormatter.format_currency(ee_change, include_sign=True) if ee_change is not None else 'N/A',
                    'Total Change': DataFormatter.format_currency(total_change, include_sign=True) if total_change is not None else 'N/A',
                })

            if comparison_rows:
                comparison_df = pd.DataFrame(comparison_rows)
                st.dataframe(comparison_df, width="stretch", hide_index=True)
                st.caption("Negative change = savings, Positive change = increase")
            else:
                st.info("No employee comparisons available")

else:
    st.subheader("💰 Contribution summary")
    st.info("""
    **No per-employee contribution data in census**

    To see current vs ICHRA cost comparison:
    1. Add **Current EE Monthly** and **Current ER Monthly** columns to your census
    2. Re-upload on the Census Input page

    These optional columns enable:
    - Per-employee cost comparison
    - Aggregate savings analysis
    - ROI calculations
    """)

# ============================================================================
# NAVIGATION
# ============================================================================

st.markdown("---")
st.success("✅ Summary complete! Ready to **Export results** →")
st.markdown("Click **4️⃣ Export results** in the sidebar to generate reports and exports")

# Feedback button in sidebar
render_feedback_sidebar()
